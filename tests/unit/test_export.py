"""The export: the release decision, the workbook round trip, and the evidence pack.

No network and no credentials. `build_bundle` takes models and returns models, and both
renderers take a bundle and return bytes, which is the whole reason the release decision was
lifted out of them — it can be tested once and asserted identical in both formats.

The test that matters most is `TestTheReleaseMappingIsTotal::test_every_status_is_mapped`,
and it earned that on its first run. The mapping originally had two tiers -- approved by a
human, or not -- and this test refused it for a missing enum member, `DRAFTED`. Following that
up is what surfaced the real problem: `DRAFTED` is what the pipeline assigns to a *successful*
answer, so a two-tier model would have told a customer that all 189 answers were unfit to
send. Three tiers, and the test now pins all three.
"""

from __future__ import annotations

import io
from datetime import UTC, datetime

import pytest
from openpyxl import Workbook, load_workbook

from attestor_core.domain import (
    Answer,
    AnswerStatus,
    Citation,
    Confidence,
    Department,
    Question,
    Review,
    Round,
    SourceRef,
    SupportVerdict,
)
from attestor_core.domain.enums import Framework, Residency, ReviewState
from attestor_platform.export import build_bundle, build_evidence_pack, fill_workbook
from attestor_platform.export.model import ExportRow, ReleaseState, release_state
from attestor_platform.export.workbook import _verification_column

SHEET = "CAIQ v4"


def _review() -> Review:
    return Review(
        review_id="rev-test01",
        customer="Northwind Traders",
        framework=Framework.CAIQ,
        residency=Residency.EU,
        current_round=1,
        state=ReviewState.AWAITING_HUMAN,
    )


def _round() -> Round:
    return Round(
        round_id="rev-test01-r1", review_id="rev-test01", ordinal=1, state=ReviewState.DRAFTING
    )


def _question(text: str, row: int, department: Department = Department.SECURITY) -> Question:
    return Question.from_text(
        text,
        department=department,
        source_ref=SourceRef(sheet=SHEET, row=row),
        framework_hint="CC6.1",
    )


def _citation(title: str = "Access Control Policy", score: float = 0.81) -> Citation:
    return Citation(
        document_uri=f"gs://attestor-505506-corpus/{title}.md",
        document_title=title,
        section="4.2 Multi-factor authentication",
        snippet="All privileged access requires a hardware security key.",
        retrieval_score=score,
        retrieved_at=datetime(2026, 8, 17, 12, 0, tzinfo=UTC),
    )


def _answer(question: Question, status: AnswerStatus, *, cited: bool = True) -> Answer:
    return Answer(
        question_id=question.question_id,
        round_id="rev-test01-r1",
        text="Yes. Privileged access requires a hardware security key.",
        citations=[_citation()] if cited else [],
        confidence=Confidence.HIGH if cited else Confidence.LOW,
        status=status,
        authored_by="SecurityAgent",
    )


def _source_workbook(rows: list[str]) -> io.BytesIO:
    """A questionnaire in the shape `parse_xlsx` reads: header row, then questions."""
    book = Workbook()
    sheet = book.active
    assert sheet is not None
    sheet.title = SHEET
    sheet.append(["Question", "Evidence reference", "Customer notes"])
    for text in rows:
        sheet.append([text, "CC6.1", ""])
    buffer = io.BytesIO()
    book.save(buffer)
    book.close()
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------------------------


#: Which statuses may omit citations, per the `Answer` validator.
_EXEMPT = {AnswerStatus.FLAGGED_NO_EVIDENCE, AnswerStatus.QUARANTINED}

#: The expected mapping, written out rather than derived, so a change to `release_state` has
#: to be a deliberate edit here too.
_EXPECTED = {
    AnswerStatus.APPROVED: ReleaseState.APPROVED,
    AnswerStatus.DRAFT: ReleaseState.SYSTEM_BACKED,
    AnswerStatus.DRAFTED: ReleaseState.SYSTEM_BACKED,
    AnswerStatus.DELIVERED: ReleaseState.SYSTEM_BACKED,
    AnswerStatus.NEEDS_HUMAN: ReleaseState.HELD,
    AnswerStatus.FLAGGED_NO_EVIDENCE: ReleaseState.NO_EVIDENCE,
    AnswerStatus.QUARANTINED: ReleaseState.QUARANTINED,
    AnswerStatus.REJECTED: ReleaseState.REJECTED,
}


class TestTheReleaseMappingIsTotal:
    def test_every_status_is_mapped(self) -> None:
        # The whole enum, not a sample. A status added later with no mapping raises, which is
        # the point: the alternative is a new status defaulting to sendable.
        assert set(_EXPECTED) == set(AnswerStatus)
        question = _question("Do you require MFA for privileged access?", 2)
        for status, expected in _EXPECTED.items():
            answer = _answer(question, status, cited=status not in _EXEMPT)
            assert release_state(answer) is expected, f"{status.value} mapped wrongly"

    def test_only_a_human_approval_reads_as_human_approved(self) -> None:
        question = _question("Do you require MFA for privileged access?", 2)
        for status in AnswerStatus:
            answer = _answer(question, status, cited=status not in _EXEMPT)
            assert release_state(answer).human_approved is (status is AnswerStatus.APPROVED)

    def test_a_flagged_answer_is_never_sendable(self) -> None:
        question = _question("Do you hold a current ISO 27001 certificate?", 2)
        for status in (
            AnswerStatus.NEEDS_HUMAN,
            AnswerStatus.FLAGGED_NO_EVIDENCE,
            AnswerStatus.QUARANTINED,
            AnswerStatus.REJECTED,
        ):
            answer = _answer(question, status, cited=status not in _EXEMPT)
            assert release_state(answer).sendable is False

    def test_a_question_with_no_answer_is_named_as_unanswered(self) -> None:
        # Not sendable, and not silently absent either -- the row still appears.
        state = release_state(None)
        assert state is ReleaseState.UNANSWERED
        assert state.sendable is False

    def test_a_status_that_claims_support_without_citations_is_refused(self) -> None:
        """The evidence outranks the status field.

        `Answer`'s validator forbids this shape, so constructing one takes going around it --
        which is exactly the situation this branch is for. A `DRAFTED` answer carrying no
        citations is a claim with nothing behind it, and the export must not call it sendable
        because a field says so.
        """
        question = _question("Do you require MFA for privileged access?", 2)
        answer = _answer(question, AnswerStatus.DRAFTED)
        # `model_construct` skips validation, the same way a hand-edited Firestore document
        # or a future refactor would.
        bypassed = Answer.model_construct(**{**answer.model_dump(), "citations": []})
        state = release_state(bypassed)
        assert state is ReleaseState.UNSUPPORTED
        assert state.sendable is False


class TestTheBundleKeepsTheCustomersOrder:
    def test_rows_come_back_in_source_row_order_not_id_order(self) -> None:
        questions = [_question(f"Question number {n} about controls?", n) for n in (7, 3, 5)]
        bundle = build_bundle(_review(), _round(), questions, [])
        assert [r.question.source_ref.row for r in bundle.rows if r.question.source_ref] == [
            3,
            5,
            7,
        ]

    def test_a_question_without_a_source_reference_sorts_last_and_is_kept(self) -> None:
        placed = _question("A placed question about encryption?", 4)
        floating = Question.from_text("A question with no source reference at all?")
        bundle = build_bundle(_review(), _round(), [floating, placed], [])
        assert len(bundle.rows) == 2
        assert bundle.rows[-1].question.question_id == floating.question_id

    def test_counts_are_computed_from_the_rows_not_asserted(self) -> None:
        first = _question("Do you require MFA for privileged access?", 2)
        second = _question("Do you encrypt customer data at rest?", 3)
        third = _question("Do you run annual penetration tests?", 4)
        answers = [
            _answer(first, AnswerStatus.APPROVED),
            _answer(second, AnswerStatus.NEEDS_HUMAN),
        ]
        bundle = build_bundle(_review(), _round(), [first, second, third], answers)
        assert bundle.answered == 2
        assert bundle.cited == 2
        assert bundle.counts[ReleaseState.APPROVED] == 1
        assert bundle.counts[ReleaseState.HELD] == 1
        assert bundle.counts[ReleaseState.UNANSWERED] == 1
        assert bundle.sendable == 1
        assert bundle.human_approved == 1


class TestTheWorkbookIsTheCustomersOwnFile:
    def test_answers_land_on_the_row_the_question_came_from(self) -> None:
        texts = [
            "Do you require MFA for privileged access?",
            "Do you encrypt customer data at rest?",
        ]
        source = _source_workbook(texts)
        # Rows 2 and 3 in the sheet, deliberately supplied in the wrong order so a renderer
        # that trusted list position rather than `source_ref` would place them swapped.
        second = _question(texts[1], 3)
        first = _question(texts[0], 2)
        answers = [_answer(first, AnswerStatus.APPROVED), _answer(second, AnswerStatus.DRAFT)]
        bundle = build_bundle(_review(), _round(), [second, first], answers)

        result = load_workbook(io.BytesIO(fill_workbook(source, bundle)))
        sheet = result[SHEET]
        # Three original columns, so Attestor's start at D.
        assert sheet["D1"].value == "Attestor answer"
        assert sheet["E2"].value == str(ReleaseState.APPROVED)
        assert sheet["E3"].value == str(ReleaseState.SYSTEM_BACKED)
        assert "hardware security key" in str(sheet["D2"].value)
        # D answer, E release, F confidence, G department, H citation count.
        assert sheet["G2"].value == "security"
        assert sheet["H2"].value == "1"

    def test_the_customers_own_columns_are_untouched(self) -> None:
        texts = ["Do you require MFA for privileged access?"]
        source = _source_workbook(texts)
        question = _question(texts[0], 2)
        bundle = build_bundle(
            _review(), _round(), [question], [_answer(question, AnswerStatus.DRAFT)]
        )
        result = load_workbook(io.BytesIO(fill_workbook(source, bundle)))
        sheet = result[SHEET]
        assert sheet["A1"].value == "Question"
        assert sheet["A2"].value == texts[0]
        assert sheet["B2"].value == "CC6.1"

    def test_a_cover_sheet_states_the_release_rule_and_the_counts(self) -> None:
        texts = ["Do you require MFA for privileged access?"]
        source = _source_workbook(texts)
        question = _question(texts[0], 2)
        bundle = build_bundle(
            _review(), _round(), [question], [_answer(question, AnswerStatus.NEEDS_HUMAN)]
        )
        result = load_workbook(io.BytesIO(fill_workbook(source, bundle)))
        assert result.sheetnames[0] == "Attestor export"
        cover = result["Attestor export"]
        values = [str(row[0]) for row in cover.iter_rows(min_col=2, max_col=2, values_only=True)]
        joined = " ".join(values)
        assert "approved" in joined
        assert str(ReleaseState.HELD) in [
            str(row[0]) for row in cover.iter_rows(min_col=1, max_col=1, values_only=True)
        ]

    def test_a_question_whose_sheet_is_missing_goes_to_the_overflow_sheet(self) -> None:
        source = _source_workbook(["Do you require MFA for privileged access?"])
        stray = Question.from_text(
            "A question from a sheet that is not in this workbook?",
            source_ref=SourceRef(sheet="Some other tab", row=9),
        )
        bundle = build_bundle(_review(), _round(), [stray], [])
        result = load_workbook(io.BytesIO(fill_workbook(source, bundle)))
        assert "Attestor — unplaced questions" in result.sheetnames
        overflow = result["Attestor — unplaced questions"]
        assert overflow.max_row == 2  # header plus the one stray question

    def test_a_very_long_answer_is_truncated_visibly(self) -> None:
        texts = ["Do you require MFA for privileged access?"]
        source = _source_workbook(texts)
        question = _question(texts[0], 2)
        answer = _answer(question, AnswerStatus.DRAFT).model_copy(update={"text": "x" * 5000})
        bundle = build_bundle(_review(), _round(), [question], [answer])
        sheet = load_workbook(io.BytesIO(fill_workbook(source, bundle)))[SHEET]
        assert "truncated for the spreadsheet" in str(sheet["D2"].value)


class TestTheEvidencePack:
    def test_it_is_a_pdf_and_it_contains_the_answers(self) -> None:
        first = _question("Do you require MFA for privileged access?", 2)
        second = _question("Do you encrypt customer data at rest?", 3, Department.ENGINEERING)
        bundle = build_bundle(
            _review(),
            _round(),
            [first, second],
            [_answer(first, AnswerStatus.APPROVED), _answer(second, AnswerStatus.NEEDS_HUMAN)],
        )
        payload = build_evidence_pack(bundle)
        assert payload.startswith(b"%PDF-")
        assert payload.rstrip().endswith(b"%%EOF")
        # Well past a cover page with nothing on it.
        assert len(payload) > 3000

    def test_an_answer_with_no_citations_says_so_rather_than_printing_nothing(self) -> None:
        question = _question("Do you hold a current ISO 27001 certificate?", 2)
        answer = _answer(question, AnswerStatus.FLAGGED_NO_EVIDENCE, cited=False)
        bundle = build_bundle(_review(), _round(), [question], [answer])
        # Rendered without raising is the assertion that matters: a citation-free answer is
        # the branch that would otherwise index into an empty list.
        assert build_evidence_pack(bundle).startswith(b"%PDF-")

    def test_it_renders_a_round_with_no_answers_at_all(self) -> None:
        questions = [_question(f"Question {n} about your controls?", n + 1) for n in range(3)]
        bundle = build_bundle(_review(), _round(), questions, [])
        assert build_evidence_pack(bundle).startswith(b"%PDF-")


class TestFilenames:
    @pytest.mark.parametrize(
        ("customer", "expected"),
        [
            ("Northwind Traders", "attestor-northwind-traders-rev-test01-r1.xlsx"),
            ("ACME / Corp.", "attestor-acme-corp-rev-test01-r1.xlsx"),
            ("???", "attestor-review-rev-test01-r1.xlsx"),
        ],
    )
    def test_a_customer_name_becomes_a_safe_filename(self, customer: str, expected: str) -> None:
        review = _review().model_copy(update={"customer": customer})
        bundle = build_bundle(review, _round(), [], [])
        assert bundle.filename("xlsx") == expected


class TestVerificationInTheDeliverable:
    """What the customer is told about who checked an answer.

    The export is where a groundedness check either means something or does not. A verdict
    that stays in the audit trail is a fact about our infrastructure; a verdict in the
    workbook is a statement to a customer, and the two answers below are different
    assurances that must not be rendered as the same one.
    """

    def _answer(self, **overrides: object) -> Answer:
        fields: dict[str, object] = {
            "question_id": "a" * 16,
            "round_id": "rev-x-r1",
            "text": "Yes, customer data is encrypted at rest.",
            "citations": [_citation()],
            "confidence": Confidence.HIGH,
            "status": AnswerStatus.DRAFTED,
            "authored_by": "SecurityAgent",
        }
        fields.update(overrides)
        return Answer(**fields)  # type: ignore[arg-type]

    def test_an_ungrounded_answer_is_not_sendable(self) -> None:
        answer = self._answer(support=SupportVerdict.UNSUPPORTED, verified_by="attestor-verifier")
        state = release_state(answer)
        assert state is ReleaseState.UNGROUNDED
        assert not state.sendable

    def test_a_human_approval_outranks_the_verifier(self) -> None:
        """A named person took responsibility, and the verdict is why they were asked.

        Letting a model's judgement override a human's signature would make the approval
        queue pointless: the operator would approve an answer and the export would refuse
        it anyway, with no way to resolve the disagreement.
        """
        answer = self._answer(
            status=AnswerStatus.APPROVED,
            support=SupportVerdict.UNSUPPORTED,
            verified_by="attestor-verifier",
        )
        state = release_state(answer)
        assert state is ReleaseState.APPROVED
        assert state.sendable

    def test_an_unverified_answer_is_still_sendable_and_says_so(self) -> None:
        # UNKNOWN is the default, which every answer written before the verifier existed
        # carries. It must not read as a pass, and it must not hold back the deliverable.
        answer = self._answer()
        assert answer.support is SupportVerdict.UNKNOWN
        assert release_state(answer) is ReleaseState.SYSTEM_BACKED
        assert _verification_column(
            ExportRow(
                question=_question("Do you encrypt at rest?", 2),
                answer=answer,
                release=ReleaseState.SYSTEM_BACKED,
            )
        ).startswith("Not verified")

    def test_a_verified_answer_names_the_agent_that_checked_it(self) -> None:
        answer = self._answer(
            support=SupportVerdict.SUPPORTED, verified_by="attestor-verifier/6871..."
        )
        rendered = _verification_column(
            ExportRow(
                question=_question("Do you encrypt at rest?", 2),
                answer=answer,
                release=ReleaseState.SYSTEM_BACKED,
            )
        )
        assert "supported" in rendered
        assert "attestor-verifier/6871..." in rendered
