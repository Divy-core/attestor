#!/usr/bin/env python3
"""Drive a review from intake to delivered using nothing but Pub/Sub messages.

    PROJECT_ID=attestor-505506 uv run python tools/run_async_review.py --limit 24 --write-proof

The Phase 4 claim is about **transport**, not scale: the 312-question numbers are Phase
3's authoritative run, and this proves the same fleet advances a review because messages
are delivered rather than because something is holding an HTTP connection open. `--limit`
keeps an iteration to a few minutes and a few cents; `--limit 0` runs the full sheet.

Every message published here goes to the real `attestor.work` topic and is read back from
a real subscription. Nothing is passed in memory between stages. The dispatcher runs in
this process rather than on Cloud Run, which is the one thing that is *not* production
shape — Phase 5 deploys it — and the handlers, the claim, the join, and the ack decisions
are all the real ones.

The message trace it prints is the exit-criterion evidence: which envelope, which
partition, which dedup key, which state it moved the review to, and what it published
next.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import tempfile
import time
import uuid
from pathlib import Path
from typing import Any

from google.cloud import pubsub_v1  # type: ignore[attr-defined]

from attestor_core.domain import Review, Round
from attestor_core.domain.enums import Framework, Residency, ReviewState
from attestor_core.protocol import WorkEnvelope, WorkKind
from attestor_platform.firestore import ReviewRepository, RoundRepository
from attestor_platform.pubsub import WorkPublisher
from attestor_platform.storage import StorageClient
from dispatcher.main import dispatch_envelope

ROOT = Path(__file__).parent.parent
PROOF_DIR = ROOT / "docs" / "proof"
CLEAN = ROOT / "seed" / "questionnaires" / "clean" / "acme-vendor-review-r1.xlsx"

SUBSCRIPTION = os.environ.get("ATTESTOR_WORK_SUBSCRIPTION", "attestor.work.local")

#: How long to wait for the next message before deciding the run has stalled. Drafting a
#: department takes minutes, so this is generous; the loop exits on DELIVERED long before.
IDLE_TIMEOUT_SECONDS = 240


def stage_questionnaire(storage: StorageClient, limit: int) -> str:
    """Put the questionnaire where intake will read it from.

    Uploaded rather than referenced from disk on purpose: `intake_document` carries a
    `gcs_uri`, and a handler that could read the local filesystem would not be the
    handler Cloud Run runs.
    """
    if limit:
        from openpyxl import load_workbook

        workbook = load_workbook(CLEAN)
        sheet = workbook.active
        # +1 for the header row.
        sheet.delete_rows(limit + 2, sheet.max_row)
        local = Path(tempfile.gettempdir()) / f"attestor-slice-{limit}.xlsx"
        workbook.save(local)
        source = local
    else:
        source = CLEAN

    object_name = f"questionnaires/async-{uuid.uuid4().hex[:8]}/{source.name}"
    return storage.upload_file(object_name, str(source), bucket_suffix="uploads")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--limit", type=int, default=24, help="questions; 0 for all 312")
    parser.add_argument("--write-proof", action="store_true")
    parser.add_argument(
        "--duplicate",
        action="store_true",
        help="redeliver every message once, to prove one delivery is one transition",
    )
    args = parser.parse_args()

    if not os.environ.get("PROJECT_ID"):
        sys.exit("error: PROJECT_ID must be set")
    project = os.environ["PROJECT_ID"]

    storage = StorageClient()
    reviews = ReviewRepository()
    rounds = RoundRepository()
    publisher = WorkPublisher()

    review_id = f"rev-async-{uuid.uuid4().hex[:8]}"
    round_id = f"{review_id}-r1"
    run_id = f"run-{int(time.time())}"

    print("=" * 78)
    print("PUB/SUB-DRIVEN REVIEW")
    print("=" * 78)
    gcs_uri = stage_questionnaire(storage, args.limit)
    print(f"  questionnaire : {gcs_uri}")
    print(f"  review        : {review_id}")
    print(f"  subscription  : {SUBSCRIPTION}")
    print(f"  duplicates    : {'every message redelivered once' if args.duplicate else 'no'}")

    reviews.put(
        Review(
            review_id=review_id,
            customer="Acme Corp (async harness)",
            framework=Framework.CAIQ,
            residency=Residency.US,
            current_round=1,
            state=ReviewState.INTAKE,
        )
    )
    rounds.put(Round(round_id=round_id, review_id=review_id, ordinal=1, state=ReviewState.INTAKE))

    # The ONLY thing this process does synchronously: publish the first envelope. From
    # here the review advances because messages arrive.
    publisher.publish(
        WorkEnvelope.for_work(
            message_id=f"{run_id}-start",
            review_id=review_id,
            run_id=run_id,
            round_id=round_id,
            kind=WorkKind.INTAKE_DOCUMENT,
            payload={"gcs_uri": gcs_uri, "original_filename": CLEAN.name},
        )
    )
    print("\n  published intake_document; the harness now only pulls and dispatches\n")

    subscriber = pubsub_v1.SubscriberClient()
    path = subscriber.subscription_path(project, SUBSCRIPTION)

    trace: list[dict[str, Any]] = []
    started = time.perf_counter()
    last_message_at = time.perf_counter()
    delivered = False

    print(f"  {'#':>3}  {'kind':<18} {'part':<12} {'dedup':<17} {'result':<12} publishes")
    print(f"  {'-' * 3}  {'-' * 18} {'-' * 12} {'-' * 17} {'-' * 12} ---------")

    while not delivered:
        if time.perf_counter() - last_message_at > IDLE_TIMEOUT_SECONDS:
            print(f"\n  no message for {IDLE_TIMEOUT_SECONDS}s -- stopping")
            break

        response = subscriber.pull(request={"subscription": path, "max_messages": 1}, timeout=30)
        if not response.received_messages:
            continue

        received = response.received_messages[0]
        last_message_at = time.perf_counter()
        envelope = WorkPublisher.decode(received.message.data)

        outcome = dispatch_envelope(envelope, attempt=1)
        if args.duplicate:
            # Same envelope, straight back in. Pub/Sub would do this on an ack timeout;
            # doing it deliberately is how the idempotency claim is proven end to end.
            repeat = dispatch_envelope(envelope, attempt=2)
            outcome["duplicate_result"] = repeat["result"]

        subscriber.acknowledge(request={"subscription": path, "ack_ids": [received.ack_id]})

        entry = {
            "n": len(trace) + 1,
            "kind": envelope.kind.value,
            "partition": envelope.partition,
            "dedup_key": envelope.dedup_key,
            "result": outcome.get("result"),
            "status_code": outcome.get("status_code"),
            "state": outcome.get("state"),
            "published": outcome.get("published", []),
            "duplicate_result": outcome.get("duplicate_result"),
        }
        trace.append(entry)
        print(
            f"  {entry['n']:>3}  {entry['kind']:<18} {(envelope.partition or '-'):<12} "
            f"{envelope.dedup_key:<17} {entry['result']!s:<12} "
            f"{len(entry['published'])}"
            + (f"   [dup: {entry['duplicate_result']}]" if args.duplicate else "")
        )

        review = reviews.get(review_id)
        if review is not None and review.state is ReviewState.DELIVERED:
            delivered = True

    elapsed = time.perf_counter() - started
    final = reviews.get(review_id)
    final_state = final.state.value if final else "unknown"

    print(f"\n  messages      : {len(trace)}")
    print(f"  final state   : {final_state}")
    print(f"  wall clock    : {elapsed:.1f}s")

    duplicates_suppressed = sum(1 for t in trace if t.get("duplicate_result") == "duplicate")
    if args.duplicate:
        print(f"  duplicates suppressed: {duplicates_suppressed}/{len(trace)}")

    passed = final_state == ReviewState.DELIVERED.value
    if args.duplicate:
        passed = passed and duplicates_suppressed == len(trace)
    print(f"\n  RESULT        : {'PASS' if passed else 'FAIL'}")

    report = {
        "case": "pubsub_driven_review",
        "pass": passed,
        "review_id": review_id,
        "round_id": round_id,
        "questions": args.limit or 312,
        "gcs_uri": gcs_uri,
        "subscription": SUBSCRIPTION,
        "messages": len(trace),
        "final_state": final_state,
        "seconds": round(elapsed, 1),
        "duplicate_mode": args.duplicate,
        "duplicates_suppressed": duplicates_suppressed,
        "trace": trace,
    }
    if args.write_proof:
        PROOF_DIR.mkdir(parents=True, exist_ok=True)
        out = PROOF_DIR / "async-review-trace.json"
        out.write_text(json.dumps(report, indent=2, sort_keys=True), encoding="utf-8")
        print(f"\nwrote {out}")

    return 0 if passed else 1


if __name__ == "__main__":
    raise SystemExit(main())
