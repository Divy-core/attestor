#!/usr/bin/env python3
"""Drive a review on the deployed stack, and measure it. No dispatcher in this process.

    PROJECT_ID=attestor-505506 uv run python tools/run_deployed_review.py --limit 0 --write-proof

The difference from `tools/run_async_review.py` is the whole point: that harness *was* the
dispatcher, pulling one message at a time and running the handlers itself. This one does
exactly two things — publish the first envelope, and watch. Every stage after that runs on
Cloud Run, driven by an Eventarc push subscription, drafting on the deployed department
engines.

That also removes the defect that stalled session two's full-scale run: the pull loop
stamped its idle timer when a message *arrived* and checked it after a ten-minute dispatch,
so it declared silence and stopped without ever asking for the remaining partitions
(`docs/proof/lost-partitions-diagnosis.md`). There is no loop here to get that wrong.

## What is observed rather than asserted

Nothing here can see inside the dispatcher, which is correct — a harness that could would
not be measuring the deployed system. Everything reported is read back from what the
deployed services wrote:

* the message trace and per-stage timing, from `work_claims`
* per-partition overlap, from claim start and completion timestamps
* achieved concurrency, from the `stage_completed` audit event the handler writes
* redelivery and lease behaviour, from the claim's attempt count and worker id
* citation rate and flag counts, from the answers themselves
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import tempfile
import time
import uuid
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from attestor_core.domain import Review, Round
from attestor_core.domain.enums import Framework, Residency, ReviewState
from attestor_core.protocol import WorkEnvelope, WorkKind
from attestor_platform.firestore import (
    AnswerRepository,
    AuditEventRepository,
    ReviewRepository,
    RoundRepository,
)
from attestor_platform.pubsub import WorkPublisher
from attestor_platform.storage import StorageClient

ROOT = Path(__file__).parent.parent
PROOF_DIR = ROOT / "docs" / "proof"
CLEAN = ROOT / "seed" / "questionnaires" / "clean" / "acme-vendor-review-r1.xlsx"

#: The measured longest partition at 312 questions, in-process (ADR/ack-deadline-margin).
#: Reported alongside the deployed figure so the comparison is explicit rather than
#: left to the reader.
IN_PROCESS_LONGEST_PARTITION = 269.0
IN_PROCESS_CONCURRENCY = 7.84
ACK_DEADLINE = 600
LEASE_SECONDS = 900

#: How long to wait with no change in review state or claim count before giving up.
#: Distinct from session two's bug: this is stamped every time *anything* moves, and it is
#: evaluated against work actually observed rather than against when a message arrived.
STALL_SECONDS = 900


def stage_questionnaire(storage: StorageClient, limit: int) -> str:
    if limit:
        from openpyxl import load_workbook

        workbook = load_workbook(CLEAN)
        sheet = workbook.active
        sheet.delete_rows(limit + 2, sheet.max_row)
        local = Path(tempfile.gettempdir()) / f"attestor-slice-{limit}.xlsx"
        workbook.save(local)
        source = local
    else:
        source = CLEAN

    object_name = f"questionnaires/deployed-{uuid.uuid4().hex[:8]}/{source.name}"
    return storage.upload_file(object_name, str(source), bucket_suffix="uploads")


def _at(value: Any) -> datetime | None:
    """Parse a claim timestamp. They are stored as ISO strings, not Firestore stamps."""
    if isinstance(value, datetime):
        return value
    if isinstance(value, str) and value:
        try:
            return datetime.fromisoformat(value)
        except ValueError:
            return None
    return None


def _seconds(start: Any, end: Any) -> float | None:
    """Elapsed seconds between two claim timestamps, or None if either is missing."""
    first, last = _at(start), _at(end)
    if first is None or last is None:
        return None
    return round((last - first).total_seconds(), 1)


def read_claims(db: Any, review_id: str) -> list[dict[str, Any]]:
    rows = db.collection("work_claims").where("review_id", "==", review_id).stream()
    claims = [dict(row.to_dict() or {}, dedup_key=row.id) for row in rows]
    return sorted(claims, key=lambda c: _at(c.get("claimed_at")) or datetime.now(UTC))


def overlap_seconds(a: dict[str, Any], b: dict[str, Any]) -> float:
    """How long two partitions were genuinely in flight at the same time.

    The number that answers "is the fan-out real?" for the *deployed* system, where the
    partitions are separate Cloud Run instances rather than threads. Zero here would mean
    Pub/Sub or Cloud Run serialised them, and the partitioning bought nothing.
    """
    starts = [_at(a.get("claimed_at")), _at(b.get("claimed_at"))]
    ends = [_at(a.get("completed_at")), _at(b.get("completed_at"))]
    if not all(starts) or not all(ends):
        return 0.0
    latest_start = max(s for s in starts if s is not None)
    earliest_end = min(e for e in ends if e is not None)
    return max(0.0, round((earliest_end - latest_start).total_seconds(), 1))


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--limit", type=int, default=0, help="questions; 0 for all 312")
    parser.add_argument("--write-proof", action="store_true")
    parser.add_argument("--proof-name", default="deployed-review-312.json")
    args = parser.parse_args()

    if not os.environ.get("PROJECT_ID"):
        sys.exit("error: PROJECT_ID must be set")

    from google.cloud import firestore

    db = firestore.Client(project=os.environ["PROJECT_ID"])
    storage = StorageClient()
    reviews = ReviewRepository()
    rounds = RoundRepository()
    answers_repo = AnswerRepository()
    audit = AuditEventRepository()

    review_id = f"rev-deployed-{uuid.uuid4().hex[:8]}"
    round_id = f"{review_id}-r1"
    run_id = f"run-{int(time.time())}"

    print("=" * 78)
    print("DEPLOYED REVIEW -- Cloud Run, Eventarc push, drafting on Agent Runtime")
    print("=" * 78)
    gcs_uri = stage_questionnaire(storage, args.limit)
    print(f"  questionnaire : {gcs_uri}")
    print(f"  review        : {review_id}")
    print(f"  questions     : {args.limit or 312}")
    print(f"  ack deadline  : {ACK_DEADLINE}s   lease: {LEASE_SECONDS}s\n")

    reviews.put(
        Review(
            review_id=review_id,
            customer="Acme Corp (deployed)",
            framework=Framework.CAIQ,
            residency=Residency.US,
            current_round=1,
            state=ReviewState.INTAKE,
        )
    )
    rounds.put(Round(round_id=round_id, review_id=review_id, ordinal=1, state=ReviewState.INTAKE))

    started = time.perf_counter()
    WorkPublisher().publish(
        WorkEnvelope.for_work(
            message_id=f"{run_id}-start",
            review_id=review_id,
            run_id=run_id,
            round_id=round_id,
            kind=WorkKind.INTAKE_DOCUMENT,
            payload={"gcs_uri": gcs_uri, "original_filename": CLEAN.name},
        )
    )
    print("  published intake_document. Nothing else in this process touches the review.\n")

    print(f"  {'elapsed':>8}  {'state':<16} claims")
    print(f"  {'-' * 8}  {'-' * 16} ------")

    last_change = time.perf_counter()
    signature: tuple[Any, ...] = ()
    final_state = "unknown"

    while True:
        elapsed = time.perf_counter() - started
        review = reviews.get(review_id)
        final_state = review.state.value if review else "unknown"
        claims = read_claims(db, review_id)
        done = sum(1 for c in claims if c.get("state") == "completed")

        current = (final_state, len(claims), done)
        if current != signature:
            signature = current
            last_change = time.perf_counter()
            print(f"  {elapsed:7.0f}s  {final_state:<16} {done}/{len(claims)} complete")

        if final_state in {ReviewState.DELIVERED.value, ReviewState.AWAITING_HUMAN.value}:
            break
        if time.perf_counter() - last_change > STALL_SECONDS:
            print(f"\n  no progress for {STALL_SECONDS}s -- stopping")
            break
        time.sleep(10)

    wall = time.perf_counter() - started
    claims = read_claims(db, review_id)
    answers = answers_repo.for_round(round_id)
    events = audit.for_review(review_id, limit=1000)

    # `work_claims` does not record the partition -- the claim is keyed on the dedup key,
    # which already encodes it (ADR-0005), but does not store it as a field. The audit
    # events do carry it, so the two are joined on the dedup key rather than the claim
    # schema being changed underneath a running deployment.
    partition_of: dict[str, str] = {}
    for event in events:
        detail = event.get("detail") or {}
        key, partition = detail.get("dedup_key"), detail.get("partition")
        if key and partition:
            partition_of[str(key)] = str(partition)

    # -- the message trace ------------------------------------------------------------
    print(f"\n  {'#':>3}  {'kind':<18} {'part':<12} {'state':<10} {'sec':>7}  worker")
    print(f"  {'-' * 3}  {'-' * 18} {'-' * 12} {'-' * 10} {'-' * 7}  ------")
    trace: list[dict[str, Any]] = []
    for index, claim in enumerate(claims, start=1):
        duration = _seconds(claim.get("claimed_at"), claim.get("completed_at"))
        record = {
            "n": index,
            "kind": claim.get("kind"),
            "partition": partition_of.get(str(claim.get("dedup_key"))),
            "dedup_key": claim.get("dedup_key"),
            "state": claim.get("state"),
            "seconds": duration,
            "worker": claim.get("worker"),
            "attempts": claim.get("attempts"),
        }
        trace.append(record)
        print(
            f"  {index:>3}  {record['kind']!s:<18} {record['partition'] or '-'!s:<12} "
            f"{record['state']!s:<10} {duration if duration is not None else '-':>7}  "
            f"{str(record['worker'])[-18:]}"
        )

    # -- did the partitions actually overlap? -----------------------------------------
    drafts = [c for c in claims if c.get("kind") == "draft_answer"]
    overlaps: dict[str, float] = {}
    for i, first in enumerate(drafts):
        for second in drafts[i + 1 :]:
            left = partition_of.get(str(first.get("dedup_key")), "?")
            right = partition_of.get(str(second.get("dedup_key")), "?")
            overlaps[f"{left}|{right}"] = overlap_seconds(first, second)

    longest_partition = max(
        (_seconds(c.get("claimed_at"), c.get("completed_at")) or 0.0 for c in drafts),
        default=0.0,
    )
    distinct_workers = sorted({str(c.get("worker")) for c in drafts if c.get("worker")})

    # -- concurrency, from what the handler recorded ----------------------------------
    concurrency: dict[str, Any] = {}
    for event in events:
        detail = event.get("detail") or {}
        if detail.get("stage") == "draft_answer" and "achieved_concurrency" in detail:
            concurrency[str(detail.get("partition"))] = {
                "achieved_concurrency": detail.get("achieved_concurrency"),
                "questions": detail.get("questions"),
                "wall_seconds": detail.get("wall_seconds"),
                "latency_sum_seconds": detail.get("latency_sum_seconds"),
                "remote_calls": detail.get("remote_calls"),
            }

    cited = [a for a in answers if a.citations]
    needs_human = [a for a in answers if a.status.value == "needs_human"]
    no_evidence = [a for a in answers if a.status.value == "flagged_no_evidence"]
    redelivered = [c for c in claims if int(c.get("attempts") or 1) > 1]

    print(f"\n  wall clock            : {wall:.1f}s")
    print(f"  final state           : {final_state}")
    print(f"  answers               : {len(answers)}")
    print(
        f"  with a citation       : {len(cited)}"
        + (f" ({100 * len(cited) / len(answers):.1f}%)" if answers else "")
    )
    print(f"  flagged for a human   : {len(needs_human)}")
    print(f"  refused for no evidence: {len(no_evidence)}")
    print(
        f"\n  longest partition     : {longest_partition:.1f}s "
        f"(in-process estimate was {IN_PROCESS_LONGEST_PARTITION}s)"
    )
    print(
        f"  margin to ack deadline: {ACK_DEADLINE - longest_partition:.1f}s "
        f"({ACK_DEADLINE / longest_partition:.1f}x)"
        if longest_partition
        else ""
    )
    print(f"  distinct workers      : {len(distinct_workers)}")
    for pair, seconds in overlaps.items():
        print(f"  overlap {pair:<28}: {seconds:.1f}s")
    for partition, stats in concurrency.items():
        print(
            f"  concurrency {partition:<12}: {stats['achieved_concurrency']} "
            f"(in-process was {IN_PROCESS_CONCURRENCY})"
        )
    print(f"  redelivered claims    : {len(redelivered)}")

    # What counts as a pass, and why it is not simply `delivered`.
    #
    # The first version of this check required `delivered`, and it called a correct run a
    # failure. A round with answers the system will not stand behind cannot close: `close_round`
    # writes commitments, and committing to an answer no human has approved is precisely what
    # the human gate exists to prevent. So a run that flags anything ends in `awaiting_human`
    # BY DESIGN, and demanding `delivered` demands that the escalation rule never fire.
    #
    # `delivered` is reachable and is the right assertion for a review where nothing was
    # flagged. Both are successes; neither is the only shape a success takes.
    #
    # What must be true either way: every stage that was reachable completed, and 60 answers
    # were persisted under the round. A run that reaches `awaiting_human` with no answers is a
    # failure wearing the same state name, which is the case this must not pass.
    stages_ok = all(str(c.get("state")) == "completed" for c in claims)
    reached_terminal = final_state in {
        ReviewState.DELIVERED.value,
        ReviewState.AWAITING_HUMAN.value,
    }
    requested = args.limit or 312
    passed = stages_ok and reached_terminal and len(answers) == requested
    print(f"\n  RESULT : {'PASS' if passed else 'FAIL'}")
    if final_state == ReviewState.AWAITING_HUMAN.value:
        print(
            f"  ({len(needs_human)} answers are held for a human, so the round cannot close. "
            "close_round follows the approvals.)"
        )

    report = {
        "case": "deployed_pubsub_review",
        "pass": passed,
        "review_id": review_id,
        "round_id": round_id,
        "run_id": run_id,
        "questions_requested": args.limit or 312,
        "gcs_uri": gcs_uri,
        "wall_seconds": round(wall, 1),
        "final_state": final_state,
        "execution": {
            "dispatcher": "Cloud Run (attestor-dispatcher), Eventarc push subscription",
            "drafting": "Agent Runtime department engines",
            "harness_role": "published one envelope, then observed",
        },
        "answers": {
            "total": len(answers),
            "with_citation": len(cited),
            "citation_rate": round(len(cited) / len(answers), 4) if answers else None,
            "needs_human": len(needs_human),
            "flagged_no_evidence": len(no_evidence),
        },
        "partitions": {
            "longest_seconds": longest_partition,
            "in_process_estimate_seconds": IN_PROCESS_LONGEST_PARTITION,
            "ack_deadline_seconds": ACK_DEADLINE,
            "lease_seconds": LEASE_SECONDS,
            "margin_seconds": round(ACK_DEADLINE - longest_partition, 1),
            "overlap_seconds": overlaps,
            "distinct_workers": distinct_workers,
        },
        "concurrency": {
            "in_process_reference": IN_PROCESS_CONCURRENCY,
            "measured": concurrency,
        },
        "redelivered_claims": [
            {
                "dedup_key": c.get("dedup_key"),
                "attempts": c.get("attempts"),
                "state": c.get("state"),
            }
            for c in redelivered
        ],
        "trace": trace,
    }
    if args.write_proof:
        PROOF_DIR.mkdir(parents=True, exist_ok=True)
        out = PROOF_DIR / args.proof_name
        out.write_text(json.dumps(report, indent=2, sort_keys=True, default=str), encoding="utf-8")
        print(f"\nwrote {out}")

    return 0 if passed else 1


if __name__ == "__main__":
    raise SystemExit(main())
