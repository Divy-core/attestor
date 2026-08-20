"""Write answers back into the customer's own workbook.

The alternative — emitting a fresh spreadsheet of our own design — is easier and worse. A
security reviewer sent a specific file with specific rows in a specific order, often with
their own tracking columns already in it, and what they want back is *that file* with the
answer column filled. Anything else is a reconciliation exercise handed to the person the
system was supposed to help.

## How a question finds its way home

``Question.source_ref`` carries the sheet name and the 1-based row the question was parsed
from, and it is preserved through triage, drafting and Firestore precisely so this is
possible. Intake records it (``agents/intake.py``), the domain model keeps it frozen, and
this module reads it back. No fuzzy text matching against cell contents, which would
mis-place a row the moment two questions in a questionnaire were worded similarly — and
real questionnaires repeat themselves constantly.

A question whose source reference is missing or whose sheet is gone from the workbook is
written to an overflow sheet rather than dropped. It has never happened in a measured run;
it is handled because losing a customer's question silently is the worst available outcome.
"""

from __future__ import annotations

import io
import logging
from pathlib import Path
from typing import IO, TYPE_CHECKING, Any

from attestor_core.domain import SupportVerdict
from attestor_platform.export.model import (
    RELEASE_RULE,
    ExportBundle,
    ExportRow,
    ReleaseState,
)

if TYPE_CHECKING:  # pragma: no cover - typing only
    from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

#: The columns Attestor appends, in order. Named for a reader rather than for a schema.
APPENDED_HEADERS = (
    "Attestor answer",
    "Release status",
    "Confidence",
    "Owning department",
    "Citations",
    "Evidence (document — section — relevance)",
    # Who checked the answer against those passages, and what they found. A column rather
    # than a footnote, because "a separate agent read this and agreed" is the strongest
    # thing this file can say about an answer no human reviewed -- and its absence, on the
    # rows where nobody checked, is the honest thing to say about those.
    "Verified by",
)

#: Row 1 is the header row, matching what ``parse_xlsx`` assumes when it reads the file.
HEADER_ROW = 1

#: Fills for the release column. Colour carries one meaning here — which of the three tiers
#: this answer is in — and nothing else in the sheet is coloured, so it cannot be mistaken for
#: anything else. Excel wants ``AARRGGBB``; these are the light tints of the console's own
#: state hues rather than a second palette invented for spreadsheets. The words in the cell
#: say the same thing, so a greyscale print loses nothing.
_FILLS: dict[ReleaseState, str] = {
    ReleaseState.APPROVED: "FFDCEBE2",
    ReleaseState.SYSTEM_BACKED: "FFEFF3F1",
    ReleaseState.HELD: "FFFBF0D9",
    ReleaseState.NO_EVIDENCE: "FFEDF0F4",
    ReleaseState.QUARANTINED: "FFEBE5F6",
    ReleaseState.REJECTED: "FFF8E2E6",
    ReleaseState.UNANSWERED: "FFF2F3F5",
    ReleaseState.UNSUPPORTED: "FFF8E2E6",
    ReleaseState.UNGROUNDED: "FFF8E2E6",
}

#: A cell holding more than this is truncated with a marker. Excel's hard limit is 32,767
#: characters and a silently truncated cell is a corrupted deliverable, so the truncation
#: is visible and the full text is always in the evidence pack.
MAX_CELL_CHARS = 3000


def _clip(text: str, limit: int = MAX_CELL_CHARS) -> str:
    if len(text) <= limit:
        return text
    return (
        f"{text[:limit].rstrip()}… [truncated for the spreadsheet; full text in the evidence pack]"
    )


def _evidence_column(row: ExportRow) -> str:
    """One line per citation: document, section, score. Readable in a cell."""
    if row.answer is None or not row.answer.citations:
        return ""
    lines = []
    for citation in row.answer.citations:
        section = citation.section or "—"
        lines.append(f"{citation.document_title} — {section} — {citation.retrieval_score:.2f}")
    return _clip("\n".join(lines))


def _verification_column(row: ExportRow) -> str:
    """Who checked it, and what they concluded. Empty is never rendered as a pass.

    An answer nobody verified says so. That row is still sendable -- citations, retrieval
    scores and a contradiction check all still stand behind it -- but the customer is told
    which of the two it is, because "grounded, and someone confirmed it" and "grounded, as
    far as anyone knows" are different assurances.
    """
    answer = row.answer
    if answer is None:
        return ""
    if answer.support is SupportVerdict.UNKNOWN:
        return "Not verified — no separate check was performed on this answer"
    who = answer.verified_by or "a separate agent"
    return f"{answer.support.value.replace('_', ' ')} — checked by {who}"


def _values(row: ExportRow) -> list[str]:
    answer = row.answer
    return [
        _clip(row.text),
        str(row.release),
        answer.confidence.value if answer else "",
        row.question.department.value,
        str(row.citation_count),
        _evidence_column(row),
        _verification_column(row),
    ]


def _write_cover(book: Any, bundle: ExportBundle) -> None:
    """A first sheet stating what this file is, who cleared what, and where it came from.

    Not decoration. Someone receives this spreadsheet weeks later with no memory of the
    conversation, and the first thing they need is whether they can act on it.
    """
    from openpyxl.styles import Alignment, Font

    sheet = book.create_sheet("Attestor export", 0)
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 72

    # An explicit cursor rather than `max_row`. A blank spacer row leaves `max_row`
    # unchanged, because openpyxl stores an empty string as no cell at all, so deriving the
    # next row from it silently overwrites the line above every separator.
    cursor = [0]

    def pair(label: str, value: str) -> None:
        cursor[0] += 1
        if not label and not value:
            return
        sheet.cell(row=cursor[0], column=1, value=label).font = Font(bold=True)
        cell = sheet.cell(row=cursor[0], column=2, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    pair("Customer", bundle.review.customer)
    pair("Review", bundle.review.review_id)
    pair("Round", f"{bundle.round_.ordinal} ({bundle.round_.round_id})")
    pair("Framework", bundle.review.framework.value)
    pair("Data residency", bundle.review.residency.value)
    pair("Review state", bundle.review.state.value)
    pair("Questions", str(len(bundle.rows)))
    pair("Answered", str(bundle.answered))
    pair("With citations", str(bundle.cited))
    pair("Sendable", str(bundle.sendable))
    pair("Approved by a human", str(bundle.human_approved))
    pair("Generated", bundle.generated_at.isoformat(timespec="seconds"))
    if bundle.origin:
        pair("Produced by", bundle.origin)

    pair("", "")
    pair("Release rule", RELEASE_RULE)
    pair("", "")
    for state, count in sorted(bundle.counts.items(), key=lambda item: -item[1]):
        pair(str(state), str(count))


def _overflow_sheet(book: Any) -> Worksheet:
    sheet = book.create_sheet("Attestor — unplaced questions")
    sheet.append(["Question", *APPENDED_HEADERS])
    return sheet


def fill_workbook(source: Path | str | IO[bytes], bundle: ExportBundle) -> bytes:
    """Return the customer's workbook with Attestor's columns filled in.

    Args:
        source: The original uploaded questionnaire — a local path, or any readable binary
            stream. The source is not modified; the result is a new workbook in memory.
        bundle: Questions joined to answers, already ordered and release-decided.

    Returns:
        The .xlsx bytes.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    # `data_only=False` keeps the customer's formulas as formulas. Reading with
    # `data_only=True` and writing back would replace every formula in their file with
    # whatever value Excel last cached, which is a destructive edit to their document.
    book = load_workbook(Path(source) if isinstance(source, (str, Path)) else source)

    # The insertion point per sheet, captured BEFORE anything is written: `max_column`
    # grows as we write, so computing it lazily inside the loop would stagger the columns
    # one step further right on every row.
    start_column: dict[str, int] = {
        sheet.title: (sheet.max_column or 0) + 1 for sheet in book.worksheets
    }
    header_written: set[str] = set()
    overflow: Worksheet | None = None
    placed = 0

    for row in bundle.rows:
        reference = row.question.source_ref
        sheet_name = reference.sheet if reference else None
        target = book[sheet_name] if sheet_name and sheet_name in book.sheetnames else None

        if target is None or reference is None or reference.row is None:
            if overflow is None:
                overflow = _overflow_sheet(book)
            overflow.append([row.question.text, *_values(row)])
            logger.warning(
                "question %s has no usable source reference; written to the overflow sheet",
                row.question.question_id,
            )
            continue

        base = start_column[target.title]
        if target.title not in header_written:
            for offset, header in enumerate(APPENDED_HEADERS):
                cell = target.cell(row=HEADER_ROW, column=base + offset, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(vertical="bottom", wrap_text=True)
            header_written.add(target.title)

        for offset, value in enumerate(_values(row)):
            cell = target.cell(row=reference.row, column=base + offset, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        release_cell = target.cell(row=reference.row, column=base + 1)
        release_cell.fill = PatternFill("solid", fgColor=_FILLS[row.release])
        placed += 1

    # Widths on the appended columns only. The customer's own columns are left exactly as
    # they set them.
    for title in header_written:
        sheet = book[title]
        base = start_column[title]
        for offset, width in enumerate((80, 30, 14, 18, 10, 52)):
            letter = sheet.cell(row=HEADER_ROW, column=base + offset).column_letter
            sheet.column_dimensions[letter].width = width

    _write_cover(book, bundle)

    buffer = io.BytesIO()
    book.save(buffer)
    book.close()
    logger.info(
        "workbook export: %d of %d rows placed by source reference", placed, len(bundle.rows)
    )
    return buffer.getvalue()
