"""Intake: an uploaded questionnaire becomes normalised `Question` records.

XLSX is parsed deterministically rather than by a model. That is a deliberate choice:
a spreadsheet has structure, and asking a model to read cells it can already read
exactly would add cost, latency, and a chance of transcription error for no benefit.
The multimodal path (the reasoning tier, see `attestor_platform.config`) is reserved
for formats that genuinely need it -- PDF, DOCX, a photograph of a printed
questionnaire -- where 3.7's document-processing gains actually apply.

`raw_text` is preserved verbatim, including hidden padding and invisible characters,
because Model Armor screens the raw cell and an injection may live precisely in what
normalisation would strip.
"""

from __future__ import annotations

import logging
from pathlib import Path

from attestor_core.domain import Question, SourceRef

logger = logging.getLogger(__name__)

#: Header names we recognise for the question column, lowercased.
_QUESTION_HEADERS = ("question", "requirement", "control", "query", "item")
#: Header names for the framework reference column.
_REF_HEADERS = ("evidence reference", "reference", "ref", "control id", "framework")

#: A cell shorter than this is a section label or a stray note, not a question.
MIN_QUESTION_CHARS = 12


def _header_index(headers: list[str], candidates: tuple[str, ...]) -> int | None:
    for index, header in enumerate(headers):
        if header.strip().lower() in candidates:
            return index
    return None


def parse_xlsx(path: str | Path, sheet: str | None = None) -> list[Question]:
    """Parse a questionnaire spreadsheet into `Question` records.

    Column detection is by header name, falling back to the widest text column, because
    real questionnaires arrive with wildly inconsistent layouts.

    Returns:
        Questions with content-derived IDs. Duplicates collapse naturally: the same
        question asked twice in one sheet yields one id, which is correct -- it is one
        question.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    worksheet = workbook[sheet] if sheet else workbook.active
    if worksheet is None:  # pragma: no cover - defensive
        return []

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(cell) if cell is not None else "" for cell in rows[0]]
    question_column = _header_index(headers, _QUESTION_HEADERS)
    ref_column = _header_index(headers, _REF_HEADERS)

    if question_column is None:
        # No recognisable header: pick the column with the most long text cells.
        widths = [
            sum(
                1
                for row in rows[1:]
                if isinstance(row[i], str) and len(row[i]) > MIN_QUESTION_CHARS
            )
            for i in range(len(headers))
        ]
        question_column = widths.index(max(widths)) if widths else 0
        logger.info("no question header found; using column %d by text density", question_column)

    questions: list[Question] = []
    seen: set[str] = set()

    for row_number, row in enumerate(rows[1:], start=2):
        if question_column >= len(row):
            continue
        cell = row[question_column]
        if not isinstance(cell, str):
            continue

        raw = cell
        # Normalised text for identity and for the model; raw_text keeps the original,
        # including the whitespace an injection may be hiding behind.
        text = " ".join(raw.split())
        if len(text) < MIN_QUESTION_CHARS:
            continue

        framework_hint = None
        if ref_column is not None and ref_column < len(row):
            value = row[ref_column]
            framework_hint = str(value).strip() if value is not None else None

        question = Question.from_text(
            raw,
            text=text,
            source_ref=SourceRef(sheet=worksheet.title, row=row_number),
            framework_hint=framework_hint or None,
        )
        if question.question_id in seen:
            continue
        seen.add(question.question_id)
        questions.append(question)

    workbook.close()
    logger.info("parsed %d questions from %s", len(questions), path)
    return questions
